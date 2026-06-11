import streamlit as st
import pandas as pd
from streamlit_gsheets import GSheetsConnection
from datetime import datetime
import io
import math

# Import thư viện xử lý Excel chuẩn form của anh
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# --- CẤU HÌNH CỐ ĐỊNH ---
COMPANY_NAME = "CÔNG TY TNHH DAYLIGHT VIỆT NAM"
COMPANY_MST = "2301380133"
COMPANY_ADDR = "Đông Lâu, Đại Đồng, Tiên Du, Bắc Ninh"
BANK_NAME_BENEFICIARY = "CÔNG TY TNHH DAYLIGHT VIỆT NAM"
BANK_STK = "Số tài khoản: 688 608 632 999"
BANK_BRANCH = "Ngân hàng: TMCP Công thương Việt Nam (Vietinbank) - CN KCN Tiên Sơn"

THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# --- CẤU HÌNH TRANG ---
st.set_page_config(page_title="DAYLIGHT VIETNAM ERP", layout="centered")

st.markdown("""
    <style>
    .main { background-color: #F1F5F9; }
    .stButton>button { width: 100%; border-radius: 10px; height: 3em; background-color: #1E3A8A; color: white; font-weight: bold; }
    .stTabs [data-baseweb="tab"] { font-weight: bold; font-size: 16px; }
    </style>
    """, unsafe_allow_html=True)

st.title("☀️ DAYLIGHT VIETNAM")

# ================= KẾT NỐI GOOGLE SHEETS =================
try:
    conn = st.connection("gsheets", type=GSheetsConnection)
    df_kho = conn.read(worksheet="KHO", ttl="1m").dropna(how="all")
except Exception as e:
    st.error(f"Chưa kết nối được Google Drive: {e}")
    df_kho = pd.DataFrame(columns=["Tên sản phẩm", "ĐVT", "Tồn", "Giá"])

if 'quote' not in st.session_state:
    st.session_state.quote = []

tab1, tab2, tab3 = st.tabs(["📦 KHO HÀNG", "📄 LẬP BÁO GIÁ", "🤝 HỢP ĐỒNG"])

# ================= TAB 1: QUẢN LÝ KHO =================
with tab1:
    st.subheader("Trạng thái tồn kho thực tế")
    if not df_kho.empty:
        st.dataframe(df_kho, use_container_width=True, hide_index=True)
    else:
        st.info("Kho hàng trống hoặc đang chờ kết nối dữ liệu...")

    with st.expander("➕ Nhập thêm vật tư / Thiết bị"):
        name = st.text_input("Tên thiết bị")
        u = st.text_input("Đơn vị tính (ĐVT)", value="Cái")
        q = st.number_input("Số lượng nhập", min_value=1, value=1)
        p = st.number_input("Giá gốc (VNĐ)", min_value=0, step=1000)

        if st.button("Xác nhận ghi vào Google Drive"):
            if name:
                new_row = pd.DataFrame([{"Tên sản phẩm": name, "ĐVT": u, "Tồn": q, "Giá": p}])
                updated_df = pd.concat([df_kho, new_row], ignore_index=True)
                conn.update(worksheet="KHO", data=updated_df)
                st.cache_data.clear()
                st.success(f"Đã đồng bộ sản phẩm '{name}' lên kho!")
                st.rerun()
            else:
                st.warning("Vui lòng nhập tên sản phẩm!")

# ================= TAB 2: LẬP BÁO GIÁ =================
with tab2:
    st.subheader("Thông tin khách hàng")
    c_name = st.text_input("Tên khách hàng / Đơn vị")
    c_phone = st.text_input("Số điện thoại")
    c_addr = st.text_input("Địa chỉ khách hàng")

    st.subheader("Chọn vật tư xuất bán")
    if not df_kho.empty:
        list_sp = df_kho["Tên sản phẩm"].tolist()
        sp_selected = st.selectbox("Sản phẩm trong kho", list_sp)

        col1, col2 = st.columns(2)
        with col1:
            sl = st.number_input("SL bán", min_value=1, value=1)
        with col2:
            vat = st.selectbox("VAT", ["0%", "5%", "8%", "10%"])

        if st.button("➕ Thêm vào danh sách"):
            row_sp = df_kho[df_kho["Tên sản phẩm"] == sp_selected].iloc[0]
            gia_goc = float(row_sp["Giá"])
            thanh_tien = sl * gia_goc * (1 + int(vat.replace("%",""))/100)

            st.session_state.quote.append({
                "Sản phẩm": sp_selected, "SL": sl, "Đơn giá": gia_goc, "VAT": vat, "Thành tiền": thanh_tien
            })
            st.toast("Đã thêm vào giỏ hàng!")
    else:
        st.warning("Không có sản phẩm nào trong kho.")

    if st.session_state.quote:
        st.markdown("---")
        df_q = pd.DataFrame(st.session_state.quote)
        st.dataframe(df_q, use_container_width=True, hide_index=True)
        tong = sum(item["Thành tiền"] for item in st.session_state.quote)
        st.error(f"TỔNG CỘNG THANH TOÁN: {tong:,.0f} VNĐ")

        # --- CHỨC NĂNG TẠO FILE EXCEL CHUẨN ---
        def generate_excel_file():
            wb = Workbook()
            ws = wb.active
            ws.title = "Báo Giá"
            
            # Setup cột
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 38
            ws.column_dimensions['C'].width = 9
            ws.column_dimensions['D'].width = 7
            ws.column_dimensions['E'].width = 14
            ws.column_dimensions['F'].width = 11 
            ws.column_dimensions['G'].width = 16 

            # Header Công ty
            ws['A1'] = COMPANY_NAME; ws['A1'].font = Font(bold=True, size=12, color="1E3A8A")
            ws['A2'] = f"Địa chỉ: {COMPANY_ADDR}"; ws['A2'].font = Font(italic=True, size=10)
            ws['A3'] = f"MST: {COMPANY_MST}"; ws['A3'].font = Font(size=10)
            
            ws.merge_cells('A5:G5')
            ws['A5'] = "BẢNG BÁO GIÁ CHI TIẾT"; ws['A5'].font = Font(bold=True, size=18, color="1E3A8A"); ws['A5'].alignment = Alignment(horizontal='center')

            # Thông tin khách
            start_info = 7
            ws.merge_cells(f'A{start_info}:G{start_info}'); ws[f'A{start_info}'] = f"Kính gửi: {c_name.upper()}"; ws[f'A{start_info}'].font = Font(bold=True)
            ws.merge_cells(f'A{start_info+1}:G{start_info+1}'); ws[f'A{start_info+1}'] = f"SĐT: {c_phone}"
            ws.merge_cells(f'A{start_info+2}:G{start_info+2}'); ws[f'A{start_info+2}'] = f"Địa chỉ: {c_addr}"

            # Bảng dữ liệu
            table_start = 11
            headers = ["STT", "TÊN SẢN PHẨM", "ĐVT", "SL", "ĐƠN GIÁ", "THUẾ VAT", "THÀNH TIỀN"]
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=table_start, column=c, value=h)
                cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = THIN_BORDER

            curr_r = table_start + 1
            for i, r in enumerate(st.session_state.quote, 1):
                vals = [i, r["Sản phẩm"], "Cái", r["SL"], r["Đơn giá"], r["VAT"], r["Thành tiền"]]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=curr_r, column=c, value=v)
                    cell.border = THIN_BORDER
                    if c == 2: cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                    else: cell.alignment = Alignment(horizontal='center', vertical='center')
                    if c == 5 or c == 7: cell.number_format = '#,##0'
                curr_r += 1

            # Tổng cộng
            ws.merge_cells(start_row=curr_r, start_column=1, end_row=curr_r, end_column=6)
            ws.cell(row=curr_r, column=1, value="TỔNG CỘNG THANH TOÁN:").alignment = Alignment(horizontal='right')
            ws.cell(row=curr_r, column=1).font = Font(bold=True, size=11)
            cell_total = ws.cell(row=curr_r, column=7, value=tong)
            cell_total.number_format = '#,##0'; cell_total.font = Font(bold=True, size=11)
            cell_total.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell_total.border = THIN_BORDER

            # Thanh toán
            curr_r += 2
            ws.cell(row=curr_r, column=1, value="THÔNG TIN THANH TOÁN:").font = Font(bold=True, underline="single"); curr_r += 1
            ws.cell(row=curr_r, column=1, value=f"Chủ tài khoản: {BANK_NAME_BENEFICIARY}").font = Font(bold=True); curr_r += 1
            ws.cell(row=curr_r, column=1, value=BANK_STK).font = Font(bold=True); curr_r += 1
            ws.cell(row=curr_r, column=1, value=BANK_BRANCH).font = Font(bold=True)

            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            excel_data = generate_excel_file()
            ten_file = f"Bao_Gia_{c_name}.xlsx" if c_name else "Bao_Gia_Daylight.xlsx"
            st.download_button(label="📥 TẢI FILE EXCEL (CHUẨN)", data=excel_data, file_name=ten_file, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            
        with col_btn2:
            if st.button("🗑️ XÓA LÀM LẠI", use_container_width=True):
                st.session_state.quote = []
                st.rerun()

        # --- NÚT XUẤT ẢNH ZALO (HTML RENDER ĐỂ CHỤP MÀN HÌNH) ---
        if st.button("📸 HIỂN THỊ BILL CHỤP ZALO", type="primary"):
            # Render HTML Bill chuyên nghiệp
            html_items = ""
            for item in st.session_state.quote:
                html_items += f"""
                <tr>
                    <td style='padding: 8px 0; border-bottom: 1px dashed #e2e8f0; color: #1e293b;'>{item["Sản phẩm"]}</td>
                    <td style='padding: 8px 0; border-bottom: 1px dashed #e2e8f0; text-align: center;'>{item["SL"]}</td>
                    <td style='padding: 8px 0; border-bottom: 1px dashed #e2e8f0; text-align: right; color: #1e3a8a; font-weight: bold;'>{item["Thành tiền"]:,.0f}</td>
                </tr>
                """
            
            zalo_bill = f"""
            <div style="background-color: white; padding: 20px; border-radius: 15px; box-shadow: 0 4px 10px rgba(0,0,0,0.1); border: 1px solid #e2e8f0; font-family: sans-serif; color: #334155; max-width: 450px; margin: 20px auto;">
                <div style="text-align: center; border-bottom: 2px solid #1e3a8a; padding-bottom: 15px; margin-bottom: 15px;">
                    <h2 style="color: #1e3a8a; margin: 0; font-size: 22px; font-weight: 800;">DAYLIGHT VIETNAM</h2>
                    <p style="margin: 5px 0 0 0; font-size: 11px; color: #64748b;">{COMPANY_ADDR}</p>
                    <p style="margin: 2px 0 0 0; font-size: 11px; color: #64748b;">Hotline: {COMPANY_MST}</p>
                    <div style="margin-top: 15px; background-color: #f8fafc; padding: 10px; border-radius: 8px; font-size: 14px; text-align: left; border: 1px solid #e2e8f0;">
                        <b>BÁO GIÁ KHÁCH HÀNG</b><br>
                        Tên/Đơn vị: <span style="color: #0f172a; font-weight: bold;">{c_name if c_name else "Khách vãng lai"}</span><br>
                        SĐT: <span style="color: #0f172a;">{c_phone}</span><br>
                        Ngày: <span style="color: #0f172a;">{datetime.now().strftime('%d/%m/%Y')}</span>
                    </div>
                </div>
                <table style="width: 100%; font-size: 13px; border-collapse: collapse; margin-bottom: 15px;">
                    <thead>
                        <tr style="color: #64748b; border-bottom: 2px solid #e2e8f0;">
                            <th style="text-align: left; padding-bottom: 8px;">Tên sản phẩm</th>
                            <th style="text-align: center; padding-bottom: 8px;">SL</th>
                            <th style="text-align: right; padding-bottom: 8px;">Thành tiền (đ)</th>
                        </tr>
                    </thead>
                    <tbody>
                        {html_items}
                    </tbody>
                </table>
                <div style="text-align: right; font-size: 18px; color: #dc2626; font-weight: 900; margin-top: 10px; border-bottom: 2px solid #e2e8f0; padding-bottom: 15px;">
                    TỔNG: {tong:,.0f} VNĐ
                </div>
                <div style="margin-top: 15px; font-size: 12px; line-height: 1.5;">
                    <b>Thông tin chuyển khoản:</b><br>
                    - CTK: {BANK_NAME_BENEFICIARY}<br>
                    - STK: <span style="color: #1e3a8a; font-weight: bold; font-size: 14px;">688 608 632 999</span><br>
                    - Vietinbank KCN Tiên Sơn
                </div>
                <div style="text-align: center; margin-top: 20px; color: #94a3b8; font-size: 11px; font-style: italic;">
                    Xin cảm ơn quý khách đã tin tưởng Daylight Vietnam!
                </div>
            </div>
            """
            st.markdown(zalo_bill, unsafe_allow_html=True)
            st.info("💡 Mẹo: Anh hãy căn chỉnh phần Bill đẹp đẽ trên màn hình rồi chụp ảnh màn hình điện thoại (Screenshot) lại và dán thẳng vào Zalo gửi cho khách nhé!")

# ================= TAB 3: HỢP ĐỒNG =================
with tab3:
    st.subheader("Soạn hợp đồng nhanh")
    if st.button("✨ AI TỰ ĐỘNG SOẠN THẢO"):
        content = f"CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM\nĐộc lập - Tự do - Hạnh phúc\n\nKhách hàng: {c_name.upper()}\nSĐT: {c_phone}\n..."
        st.text_area("Bản xem trước nội dung", value=content, height=300)
